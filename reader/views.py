from django.http import JsonResponse
from django.shortcuts import render, redirect
from collections import defaultdict
import openpyxl
import copy
import reader.data as subject_data
from .models import Counter, Upload_user
from django.utils import timezone
from django.http import HttpResponse, HttpResponseRedirect
from datetime import datetime, timedelta

def ge_not_list(request):
	context = request.session['context']
	return JsonResponse(context['GE_not'], safe=False)

def major_req_not_list(request):
	context = request.session['context']
	return JsonResponse(context['Major_req_not'], safe=False)

def major_sub_not_list(request):
	context = request.session['context']
	return JsonResponse(context['Major_sub_not'], safe=False)

def resub_list(request):
	context = request.session['context']
	return JsonResponse(context['sorted_grade'], safe=False) 

def index(request):
	try:
		counter = Counter.objects.get(date=timezone.now())
	except Counter.DoesNotExist:
		counter = Counter.objects.create(count=0, date=timezone.now())

	if request.COOKIES.get('count') == 'visited':
		return render(request, 'reader/index.html')
	else:
		key = 'count'
		value = 'visited'
		now = datetime.now()
		end_of_day = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
		expire = end_of_day.strftime('%a, %d %b %Y %H:%M:%S GMT')

		res = HttpResponse(render(request, 'reader/index.html'))
		res.set_cookie(key, value=value, expires=expire)
		counter.count += 1
  
	counter.save()

	return res

def upload_or_result(request):
	if 'context' in request.session:
		return render(request, 'reader/result.html', request.session['context'])
	return render(request, 'reader/upload.html')

def delete_file(request):
	if 'context' in request.session:
		del request.session['context']
	return redirect('/upload')

def sort_by_grade():
	global subject_did, GE_not, no_sub, re_sub
	score_for_grade = {'A+': 4.5, 'A0': 4.0, 'B+': 3.5, 'B0':3.0, 
					'C+': 2.5, 'C0': 2.0, 'D+':1.5, 'D0': 1.0, 'P': 5, 'F': 0, 'W': 10}

	did = dict()
	for key in subject_did.keys():
		if key[1:] == '채플':
			continue
		if key[-4:] == '(재수)':
			continue
		elif key[1:] in re_sub or key in re_sub: # F인데 재수강 했음 or 그냥 재수강 했음
			continue
		else:
			if key[1:] in subject_data.GE_change.keys():
				new = key[1:]
				try:
					while subject_data.GE_change[new]:
						new = subject_data.GE_change[new]
				except:
					for key2 in subject_did.keys():
						if new == key2:
							new = 0
							break
					if new:
						did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]
			elif key in subject_data.GE_change.keys():
				new = key
				try:
					while subject_data.GE_change[new]:
						new = subject_data.GE_change[new]
				except:
					for key2 in subject_did.keys():
						if new == key2:
							new = 0
							break
					if new:
						did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]
       
			else:
				did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]
   
	value = list(did.values())
	value = sorted(value, key=lambda x: x[2])

	json_parse = dict()
	for key in value:
		# key[-1] => 과목명
		if key[-3] >= 3:
			break
		json_parse[key[-1]] = {'subject': key[-1], 'score': key[-2], 'category': subject_did[key[-1]][0]}
	
	if(len(json_parse)==0):
			json_parse["none"] = {'subject': '<p class="text-danger">재수강할 과목이 없습니다.</p>', 'score': '', 'category': ''}

	json_list = list(json_parse.values())
	return json_list

def GE_not_list():
	global student, area_did
	GE_not = copy.deepcopy(subject_data.GE_list)
	GE_did = set()
	no_com_co = {'산업경영공학과', '컴퓨터공학과', '미디어소프트웨어학과'}	

	for i in area_did['교필']:
		tmp = remove_jaesu(i[0])
		GE_did.add(copy.deepcopy(tmp))

	for sub in GE_did:
		while(sub in subject_data.GE_change):
			sub = subject_data.GE_change[sub]
		if(sub in GE_not):
			del GE_not[sub]

	if (int(student['student_num'][2:4]) < 21 or student['major'] in no_com_co) and ('컴퓨팅사고와 코딩기초' in GE_not):
		del GE_not['컴퓨팅사고와 코딩기초']
	
	if(len(GE_not)==0):
			GE_not["none"] = {'subject': '<p class="text-danger">미수강한 교필이 없습니다.</p>', 'score': '', 'category': ''}
	
	return GE_not

def remove_jaesu(text):
    if text.endswith("(재수)"):
        text = text[:-4]
    if text.startswith("★"):
        text = text[1:]
    return text

def Major_sub():
	global student, area_did

	if(not student['major'] in subject_data.CM_list):
		Major_sub_not = {}
		Major_sub_not["none"] = {'subject': '<p class="text-danger">지원하지 않는 학과입니다.</p>', 'score': '', 'category': ''}
		return Major_sub_not
	
	Major_sub_not = copy.deepcopy(subject_data.CM_list[student["major"]]["list"])
	Ms_did = set()

	for i in area_did['전필']:
		tmp = remove_jaesu(i[0])
		Ms_did.add(copy.deepcopy(tmp))
	for i in area_did['전선']:
		tmp = remove_jaesu(i[0])
		Ms_did.add(copy.deepcopy(tmp))

	for sub in Ms_did:
		while(sub in subject_data.CM_list[student["major"]]["change"]):
			sub = subject_data.CM_list[student["major"]]["change"][sub]
		if(sub in Major_sub_not):
			del Major_sub_not[sub]

	if(len(Major_sub_not)==0):
		Major_sub_not["none"] = {'subject': '<p class="text-danger">미수강한 전선이 없습니다.</p>', 'score': '', 'category': ''}
	
	return Major_sub_not

def Major_req(Major_sub_not):
	global student
	s_num = int(student['student_num'][0:4])#학번찾음
	
	Major_req_not = {}
	if(not student['major'] in subject_data.CM_list):
		Major_req_not["none"] = {'subject': '<p class="text-danger">지원하지 않는 학과입니다.</p>', 'score': '', 'category': ''}
		return Major_req_not
	Major_req = copy.deepcopy(subject_data.CM_list[student["major"]]["REQ"][s_num])
	
	for sub in Major_req:
		while(sub in subject_data.CM_list[student["major"]]["change"]):
			sub = subject_data.CM_list[student["major"]]["change"][sub]
		if(sub in Major_sub_not):
			Major_req_not[sub] = Major_sub_not[sub]
			del Major_sub_not[sub]

	if(len(Major_req_not)==0):
		Major_req_not["none"] = {'subject': '<p class="text-danger">미수강한 전필이 없습니다.</p>', 'score': '', 'category': ''}

	return Major_req_not

def area_change(Major_req_did, Major_sub_did):
	global student
	s_num = int(student['student_num'][0:4])#학번찾음
	need_change = []
	sub_did = set(remove_jaesu(item[0]) for item in Major_sub_did)
	req_did = set(remove_jaesu(item[0]) for item in Major_req_did)
	
	if(not student['major'] in subject_data.CM_list):
		return ["지원하지 않는 학과입니다."]

	Major_req = copy.deepcopy(subject_data.CM_list[student["major"]]["REQ"][s_num])

	for did in Major_req:
		if(did in sub_did):
			need_change.append({"before":did+"(전선)", "after":did+"(전필)"})
		if(did in req_did):
			req_did.remove(did)

	for did in req_did:
		need_change.append({"before":did+"(전필)", "after":did+"(전선)"})

	return need_change

def grad_cond():
	global student
	s_num = int(student['student_num'][0:4])
	major = student['major']
	if(major in subject_data.CM_list) :
		return subject_data.CM_list[major]["grad"][s_num]

	return ['지원하지 않는 학과입니다.']

def upload_file(request):
	global student, area, short_area, score_need_list, score_for_grade, info_category, year, score_need, score_did, subject_did, area_did, semester_grade, semester_subject, GE_not, no_sub, re_sub
	file = request.FILES['uploaded_file']
	if file:
		if file.name.endswith('xlsx') or file.name.endswith('xls'):
			
			wb = openpyxl.load_workbook(file)
			sheet = wb.active
		#----------------------------------------------------------------------------------------------
			# already set
			area = {'교필영역', '교선영역', '전필영역', 
					'전선영역', '복필영역', '복전영역', '부전공영역', 
					'일선영역', '교직영역'} #, '졸업사정결과'}
			short_area = {'교필', '교선', '전필', 
					'전선', '복필', '복전', '부전', 
					'일선', '교직', '채플'}
			score_need_list = {'교양요구학점', '전필요구학점', '전선요구학점', '복수전공필수요구학점', 
							'복수전공요구학점', '부전공요구학점'}
			score_for_grade = {'A+': 4.5, 'A0': 4.0, 'B+': 3.5, 'B0':3.0, 
							'C+': 2.5, 'C0': 2.0, 'D+':1.5, 'D0': 1.0}
			info_category = ['major', 'minor', 'student_num', 'grade', 
						'name', 'score_need', 'score_did', 'admsn']
			info_idx = ['A2', 'F2', 'J2', 'M2', 'O2', 'Y2', 'AB2', 'Q2']
			year = ['2015', '2016', '2017', '2018', '2019', '2020', '2021', 
				'2022', '2023', '2024', '2025', '2026']
			semester = {'1학기', '2학기'}
			season_semester = ['여름학기', '겨울학기']

			# set during runtime
			score_need = defaultdict(int)	# 영역별 요구 학점
			score_did = defaultdict(float)	# 영역별 이수 학점
			score_did_np = defaultdict(float)	# no_p
			subject_did = defaultdict(list)	# 수강한 과목 모두 (이건 하지 말까 고민중 semester_subject를 모아둔 느낌) => 필요하다!! # [영역, 학점, 등급]
			no_sub = defaultdict(list)
			subject_didnot = defaultdict(dict) # 수강하지 않은 필수 과목 => set(필수 과목) 해서 discard 하는방식
			area_did = defaultdict(list) # 영역별 [과목명, 이수학점, 등급, 년도, 학기]

			semester_grade = defaultdict(dict) # 학년별, 학기별 평점(grade), 이수 학점(score)
			for i in range(1, 6):	# 학년별
				for j in range(1, 3):	# 학기별
					semester_grade[str(i)+'_'+str(j)] = defaultdict(float)

			semester_subject = defaultdict(dict) # 학년별, 학기별 수강 과목
			for i in range(1, 6):	# 학년별
				for j in range(1, 3):	# 학기별
					semester_subject[str(i)+'_'+str(j)] = list()

			student = dict()
			for i in range(8):
				student[info_category[i]] = sheet[info_idx[i]].value
			

			ex = set()
			season_ex = set()
			j = 1
			while j < sheet.max_row + 1: # 수강 학기 매칭 ex) 1학년 1학기 => 19년도 1학기
				if sheet[str('X'+str(j))].value in year:
					if sheet[str('Z'+str(j))].value in semester:
						ex.add((str(sheet[str('X'+str(j))].value), str(sheet[str('Z'+str(j))].value)[0]))	# ex) (2019,1)
					elif sheet[str('Z'+str(j))].value in season_semester:
						season_ex.add((str(sheet[str('X'+str(j))].value), str(sheet[str('Z'+str(j))].value)))
				if sheet[str('A'+str(j))].value and sheet[str('A'+str(j))].value[:-3] in score_need_list: 
					score_need[str(sheet[str('A'+str(j))].value[:-3])] = int(str(sheet[str('N'+str(j))].value))
					score_need[str(sheet[str('P'+str(j))].value[:-3])] = int(str(sheet[str('W'+str(j))].value))
				j += 1

			student_semester = []
			for i in ex:
				student_semester.append(i)
			student_semester.sort()
		
			student_season_semester = []
			for i in season_ex:
				student_season_semester.append(i)
			student_season_semester.sort()

			semester_dict = dict()
			j = 0
			for i in range(len(student_semester)):
				j += 1
				semester_dict[student_semester[i]] = str((i//2)+1)+'_'+str(j) # ex) [('2019', '1')] = (1, 1) 
				j %= 2
			j = 0
			for i in range(len(year)):
				semester_grade[year[i]+'_'+season_semester[j%2]] = defaultdict(float)
				semester_subject[year[i]+'_'+season_semester[j%2]] = list()
				semester_dict[(year[i], season_semester[j%2])] = year[i]+'_'+season_semester[j%2] # ex) [('2019', '여름학기')] = (2019, '여름학기')
				j += 1

			j = 1
			while j < sheet.max_row + 1: # just for the example

				if sheet[str('A'+str(j))].value and sheet[str('A'+str(j))].value[:3] == '교필:':
					total_data = list(sheet[str('A'+str(j))].value.split(' '))
					for data in total_data:
						sub, score = map(str, data.split(':'))
						score_did[sub] += float(score)

				if sheet[str('A'+str(j))].value in area: # ex) 교필영역이 area집합에 있으면
					j += 2
			# 영역  개설학부(과)  이수구분	강좌명	학점  등급	년도  학기  비고
			# A		B			G		I    S	  U    X	Z	AB
			# F면 학점은 0으로, 탭에는 나오게
			# 재수강하지 않았으면 추천 목록에 띄우기

					while sheet[str('G'+str(j))].value in short_area:	# ex) 해당 영역(교필영역)에 수강한 과목이 있으면
						try:
							ex_semester = (semester_dict[(sheet[str('X'+str(j))].value, sheet[str('Z'+str(j))].value[0])])
						except:
							ex_semester = (str(sheet[str('X'+str(j))].value)+'_'+sheet[str('Z'+str(j))].value)
						if sheet[str('U'+str(j))].value != 'P' == 'W':
							continue
						if sheet[str('U'+str(j))].value != 'P' and sheet[str('U'+str(j))].value != 'F': # F맞으면 학점이 하이폰(-)으로 나오나요?
							try:
								semester_grade[ex_semester]['S'] += int(sheet[str('S'+str(j))].value) # 해당 과목 이수 학점
								semester_grade[ex_semester]['G'] += int(sheet[str('S'+str(j))].value)*score_for_grade[sheet[str('U'+str(j))].value]
							except:
								pass

						if sheet[str('U'+str(j))].value == 'P':
							semester_grade[ex_semester]['P'] += int(sheet[str('S'+str(j))].value)

						# [영역, 학점, 등급]
						
						subject_did[sheet[str('I'+str(j))].value] = [sheet[str('G'+str(j))].value, sheet[str('S'+str(j))].value, sheet[str('U'+str(j))].value]
						if sheet[str('U'+str(j))].value[-4:] == 'F' or sheet[str('S'+str(j))].value == '-':
							no_sub[sheet[str('I'+str(j))].value[1:]] = [sheet[str('G'+str(j))].value, sheet[str('S'+str(j))].value, sheet[str('U'+str(j))].value]
						
						semester_subject[ex_semester].append(sheet[str('I'+str(j))].value)
						# [과목명, 이수학점, 등급, 년도, 학기]
						area_did[sheet[str('G'+str(j))].value].append([sheet[str('I'+str(j))].value, sheet[str('S'+str(j))].value, sheet[str('U'+str(j))].value, sheet[str('X'+str(j))].value, sheet[str('Z'+str(j))].value])

						j += 1	# excel 다음 행으로
						# 	break	# 요구학점 나오면 해당 영역 종료

				else:
					j += 1

			for i in range(1, 6):
				for j in range(1, 3):
					if semester_grade[str(i)+'_'+str(j)]['S']:
						semester_grade[str(i)+'_'+str(j)]['G'] = round(semester_grade[str(i)+'_'+str(j)]['G'] / semester_grade[str(i)+'_'+str(j)]['S'], 2)
						semester_grade[str(i)+'_'+str(j)]['S'] += semester_grade[(str(i)+'_'+str(j))]['P']  # 계산할때는 패논패 계산 없이 함
		
			total_avg = sheet[str('H'+str(sheet.max_row - 1))].value # 총 평점 평균
			church = sheet[str('L'+str(sheet.max_row - 1))].value # 채플
			ratio = dict()
			try:
				ratio['전필'] = min(100.0, round(score_need['전필이수학점'] / score_need['전필요구학점'], 2)*100)
			except:
				ratio['전필'] = 0
			try:
				ratio['전선'] = min(100.0, round(score_need['전선이수학점'] / score_need['전선요구학점'], 2)*100)
			except:
				ratio['전선'] = 0
			try:
				ratio['교양'] = min(100.0, round(score_need['교양이수학점'] / score_need['교양요구학점'], 2)*100)
			except:
				ratio['교양'] = 0
   			# 교필은 아직임

			grade_key = {'A+': 'AP', 'A0': 'A', 'B+': 'BP', 'B0': 'B', 
							'C+': 'CP', 'C0': 'C', 'D+': 'DP', 'D0': 'D', 'F': 'F'}
			
			ratio['등급'] = dict()
			for key in grade_key.values():
				ratio['등급'][key] = 0
		
			re_sub = set()
			cnt = 0
			for sub in subject_did:
				if sub[-4:] == "(재수)":
					re_sub.add(sub) # 과목명(재수)
					re_sub.add(sub[:-4]) # 과목명
					try:
						tmp = {v:k for k,v in subject_data.IME_change.items()}
						if remove_jaesu(sub) in tmp.keys():
							re_sub.add(tmp.get(remove_jaesu(sub)))
						elif remove_jaesu(sub) in tmp[remove_jaesu(sub)]:
							re_sub.add(tmp.get(remove_jaesu(sub)))
					except:
						pass
					try:
						if subject_did[sub][-1] != 'P' and subject_did[sub][1] != '-':
							ratio['등급'][grade_key[subject_did[sub][-1]]] += 1
							cnt += 1
					except:
						pass

			for sub in subject_did:
				if sub[1:] not in re_sub and sub not in re_sub: # 별과목, 과목 인지 몰라욤
					try:
						if subject_did[sub][-1] != 'P' and subject_did[sub][1] != '-':
							ratio['등급'][grade_key[subject_did[sub][-1]]] += 1
							cnt += 1
					except:
						pass
					
			for key in ratio['등급'].keys():
				ratio['등급'][key] = round(ratio['등급'][key] / cnt * 100, 2)

			major_grade = defaultdict(float)

			for data in area_did['전필']:
				if data[2] == 'P' and data[2] != 'F' and data[1] != '-':
					score_did_np['전필'] += int(data[1])
				if data[2] != 'P' and data[2] != 'F' and data[1] != '-':
					major_grade['전필'] += float(score_for_grade[data[2]])*float(data[1])
					major_grade['전공'] += float(score_for_grade[data[2]])*float(data[1])
			try:
				major_grade['전필'] = round(major_grade['전필'] / (score_did['전필'] - score_did_np['전필']), 2)
			except:
				major_grade['전필'] = 0
			
			for data in area_did['전선']:
				if data[2] == 'P' and data[2] != 'F' and data[1] != '-':
					score_did_np['전선'] += int(data[1])
				if data[2] != 'P' and data[2] != 'F' and data[1] != '-':
					try:
						major_grade['전선'] += float(score_for_grade[data[2]])*float(data[1])
						major_grade['전공'] += float(score_for_grade[data[2]])*float(data[1])
					except:
						pass
			try:
				major_grade['전선'] = round(major_grade['전선'] / (score_did['전선'] - score_did_np['전선']), 2)
			except:
				major_grade['전선'] = 0
    

			for data in area_did['부전']:
				if data[2] == 'P' and data[2] != 'F' and data[1] != '-':
					score_did_np['부전'] += int(data[1])
				if data[2] != 'P' and data[2] != 'F' and data[1] != '-':
					try:
						major_grade['부전'] += float(score_for_grade[data[2]])*float(data[1])
						major_grade['전공'] += float(score_for_grade[data[2]])*float(data[1])
					except:
						pass
			try:
				major_grade['부전'] = round(major_grade['부전'] / (score_did['부전'] - score_did_np['부전']), 2)
			except:
				major_grade['부전'] = 0


			for data in area_did['복전']:
				if data[2] == 'P' and data[2] != 'F' and data[1] != '-':
					score_did_np['복전'] += int(data[1])
				if data[2] != 'P' and data[2] != 'F' and data[1] != '-':
					try:
						major_grade['복전'] += float(score_for_grade[data[2]])*float(data[1])
						major_grade['전공'] += float(score_for_grade[data[2]])*float(data[1])
					except:
						pass
			try:
				major_grade['복전'] = round(major_grade['복전'] / (score_did['복전'] - score_did_np['복전']), 2)
			except:
				major_grade['복전'] = 0
    
			# 전필, 전선, 복전, 부전
			try:
				major_grade['전공']	= round(major_grade['전공'] / (score_did['전선']+score_did['전필']+score_did['복전']+score_did['부전']-score_did_np['전선']-score_did_np['전필']-score_did_np['복전']-score_did_np['부전']), 2)
			except:
				major_grade['전공'] = 0
    
			sorted_subject = [] # 성적순으로 정렬된것 만들어야함
		
			GE_not = list(GE_not_list().values())
			Major_sub_not = Major_sub()
			Major_req_not = list(Major_req(Major_sub_not).values())
			Major_sub_not = list(Major_sub_not.values())
			need_change = area_change(area_did['전필'], area_did['전선'])
			sorted_grade = sort_by_grade() # 재수강 추천

			context = {'area_did': area_did, 
					'semester_grade': semester_grade,
					'semester_subject': semester_subject,
					'student': student,
					'subject_did': subject_did,
					'score_did': score_did,
					'score_need': score_need,
					'total_avg': total_avg,
					'church': church,
					'ratio': ratio,
					'major_grade': major_grade,
					'sorted_subject': sorted_subject,
					'GE_not': GE_not,
					'Major_sub_not' : Major_sub_not,
					'Major_req_not' : Major_req_not,
					'grad' : grad_cond(),
					'sorted_grade': sorted_grade, # 이수한 과목 중 성적 낮은것부터 리스트로
					'need_change': need_change,
					'no_sub': no_sub,
			}
			request.session["context"] = context
		else:
			return redirect('/upload')
	for i in context:
		print(i," ",context[i])
		print()
	
	try:
		user = Upload_user.objects.get(date=timezone.now())
	except Upload_user.DoesNotExist:
		user = Upload_user.objects.create(count=0, date=timezone.now())

	if request.COOKIES.get('upload') == 'thankyou':
		return redirect('/upload')
	else:
		key = 'upload'
		value = 'thankyou'
		now = datetime.now()
		end_of_day = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
		expire = end_of_day.strftime('%a, %d %b %Y %H:%M:%S GMT')

		res = HttpResponseRedirect('/upload')
		res.set_cookie(key, value=value, expires=expire)
		user.count += 1
		user.save()
  
		return res
